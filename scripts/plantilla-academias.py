"""Plantilla de Excel para que el club cargue los inscritos de las academias.

Una fila por NIÑO y por DÍA que asiste. Las columnas son, a propósito, el modelo
de datos: si una columna no cabe, el modelo está mal.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

SALIDA = "/Users/laurasalazar/Documents/Proyectos/Centro de Control CDAF/Centro de Control CDAF/plantillas/academias-inscritos.xlsx"

# ── Listas cerradas (salen de la base de datos real) ──
DEPORTES = ["Tenis", "Pádel"]
ACADEMIAS = ["Recreativa", "Competencia"]
NIVELES = ["Bola Roja", "Bola Naranja", "Bola Verde", "Bola Amarilla",
           "Principiantes", "Iniciados", "Intermedio"]
DIAS = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
DURACIONES = [60, 90, 120]
PROFESORES = ["Cristian Castro", "Dairon Guarín", "Esteban Graciano", "Joaquín",
              "Jorge", "Leo Ruíz", "Sebastián Niño", "Willington"]
CANCHAS = [1, 2, 3, 4]

COLUMNAS = [
    ("Nombres",        18, "Nombre del NIÑO que asiste, no del papá ni de la mamá."),
    ("Apellidos",      18, "Apellidos del niño."),
    ("Documento",      14, "Cédula o tarjeta de identidad del niño. Si no lo tienes, déjalo vacío: lo cruzamos por nombre, pero con documento el cruce es exacto."),
    ("Deporte",        10, "Tenis o Pádel. Se escoge de la lista."),
    ("Academia",       14, "Recreativa o Competencia. Se escoge de la lista. Junto con el deporte define a cuál de las 4 academias entra."),
    ("Nivel",          15, "Nivel del NIÑO. Se escoge de la lista para que los reportes por nivel sirvan."),
    ("Día",             8, "Día de la semana en que viene. UNA FILA POR CADA DÍA."),
    ("Hora",           10, "Hora en que empieza la clase ese día. Formato 24 horas: 16:30, no 4:30 pm."),
    ("Duración (min)", 14, "Cuánto dura la clase: 60, 90 o 120 minutos."),
    ("Profesor",       19, "Quién le da la clase ESE día. Puede ser distinto en otro día."),
    ("Cancha",          9, "Número de cancha de ese día."),
]

# El caso que contó el coordinador: el mismo niño, dos profesores distintos.
EJEMPLOS = [
    ["Pepito (EJEMPLO)",  "Pérez",  "1098765432", "Tenis", "Recreativa", "Bola Verde", "Mar", "16:30", 60, "Jorge",            3],
    ["Pepito (EJEMPLO)",  "Pérez",  "1098765432", "Tenis", "Recreativa", "Bola Verde", "Jue", "16:30", 60, "Jorge",            3],
    ["Pepito (EJEMPLO)",  "Pérez",  "1098765432", "Tenis", "Recreativa", "Bola Verde", "Sáb", "12:00", 60, "Esteban Graciano", 4],
    ["Juanita (EJEMPLO)", "Gómez",  "1076543210", "Pádel", "Competencia", "Intermedio", "Lun", "07:00", 90, "Willington",      1],
]

ARIAL = "Arial"
AZUL = "1F3A4D"
LIMA = "D4E157"
AMARILLO = "FFF9C4"
GRIS = "F5F5F5"

wb = Workbook()

# ═══════════════════ Hoja 1: Instrucciones ═══════════════════
ins = wb.active
ins.title = "Instrucciones"
ins.sheet_view.showGridLines = False
ins.column_dimensions["A"].width = 3
ins.column_dimensions["B"].width = 104

def linea(fila, texto, size=11, bold=False, color="000000", espacio_arriba=0):
    if espacio_arriba:
        ins.row_dimensions[fila].height = espacio_arriba
    c = ins.cell(row=fila, column=2, value=texto)
    c.font = Font(name=ARIAL, size=size, bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return c

f = 2
linea(f, "ACADEMIAS · Listado de inscritos", size=18, bold=True, color=AZUL); f += 1
linea(f, "Centro Deportivo Alejandro Falla", size=11, color="666666"); f += 2

linea(f, "Qué hay que llenar", size=13, bold=True, color=AZUL); f += 1
linea(f, "La hoja «Inscritos». Una fila por cada niño y por cada día que viene.", bold=True); f += 1
linea(f, "Si un niño viene tres veces a la semana, son tres filas con su nombre repetido. Eso es a propósito: "
         "así se puede decir que el martes lo atiende un profesor y el sábado otro, en otra cancha y a otra hora.")
ins.row_dimensions[f].height = 30; f += 2

linea(f, "Ejemplo: Pepito viene martes y jueves a las 4:30 p.m. con Jorge, y los sábados a las 12 con Graciano", size=11, bold=True); f += 1
ej = [
    ["Nombres", "Apellidos", "Día", "Hora", "Profesor", "Cancha"],
    ["Pepito", "Pérez", "Mar", "16:30", "Jorge", "3"],
    ["Pepito", "Pérez", "Jue", "16:30", "Jorge", "3"],
    ["Pepito", "Pérez", "Sáb", "12:00", "Esteban Graciano", "4"],
]
for i, fila_ej in enumerate(ej):
    texto = "     " + "  ·  ".join(str(x).ljust(9) for x in fila_ej)
    c = linea(f, texto, size=10, bold=(i == 0), color=("444444" if i == 0 else "000000"))
    c.font = Font(name="Courier New", size=10, bold=(i == 0), color=("444444" if i == 0 else "000000"))
    f += 1
f += 1
linea(f, "Son tres filas del MISMO niño. Una sola inscripción, con tres horarios.", size=10, color="666666"); f += 2

linea(f, "Reglas importantes", size=13, bold=True, color=AZUL); f += 1
for regla in [
    "El nombre es del NIÑO que entrena, no del papá ni de quien paga.",
    "Un niño no puede estar en Recreativa y en Competencia del mismo deporte. Sí puede hacer tenis y pádel.",
    "Las columnas con lista desplegable SOLO aceptan las opciones de la lista. Es para que el mismo profesor "
    "no quede escrito de tres formas distintas y los reportes sirvan.",
    "La hora va en formato de 24 horas: 16:30, no 4:30 pm.",
    "No hay que poner el precio ni la matrícula: la plata sale de las facturas de Siigo.",
    "No hay que agrupar nada ni poner nombres de grupo. El sistema arma los grupos solo, cruzando día + hora + profesor.",
]:
    c = linea(f, "•  " + regla)
    ins.row_dimensions[f].height = 28 if len(regla) > 95 else 15
    f += 1
f += 1

linea(f, "Antes de enviarlo", size=13, bold=True, color=AZUL); f += 1
c = linea(f, "Borra las filas amarillas de ejemplo (las que dicen EJEMPLO).", bold=True, color="B71C1C"); f += 1
linea(f, "Si un dato no lo tienes, déjalo vacío antes que inventarlo. Un documento equivocado le atribuye la "
         "asistencia a otra persona.")
ins.row_dimensions[f].height = 28; f += 2

linea(f, "Qué significa cada columna", size=13, bold=True, color=AZUL); f += 1
for nombre, _, ayuda in COLUMNAS:
    c = linea(f, f"{nombre}  —  {ayuda}", size=10)
    ins.row_dimensions[f].height = 26 if len(ayuda) > 90 else 14
    f += 1

# ═══════════════════ Hoja 3 (creada antes para referenciarla): Listas ═══════════════════
listas = wb.create_sheet("Listas")
listas.sheet_state = "visible"  # visible a propósito: si el club necesita otra opción, se ve dónde pedirla
listas["A1"] = "Opciones válidas de cada columna, para consulta. Si falta alguna, pídenosla."
listas["A1"].font = Font(name=ARIAL, size=10, bold=True, color="B71C1C")

BLOQUES = [
    ("Deporte", DEPORTES), ("Academia", ACADEMIAS), ("Nivel", NIVELES),
    ("Día", DIAS), ("Duración", DURACIONES), ("Profesor", PROFESORES), ("Cancha", CANCHAS),
]
rangos = {}
for col_idx, (titulo, valores) in enumerate(BLOQUES, start=1):
    L = get_column_letter(col_idx)
    listas.column_dimensions[L].width = max(12, len(titulo) + 4)
    c = listas.cell(row=3, column=col_idx, value=titulo)
    c.font = Font(name=ARIAL, size=10, bold=True)
    for i, v in enumerate(valores, start=4):
        vc = listas.cell(row=i, column=col_idx, value=v)
        vc.font = Font(name=ARIAL, size=10)
    # Lista EN LÍNEA, no referencia a rango. Dos razones:
    #   1. Google Sheets importa las listas en línea sin falla; las que apuntan a
    #      otra hoja se le pierden a veces y los desplegables salen mudos.
    #   2. El estándar OOXML pide formula1 SIN el "=" delante. Excel lo perdona,
    #      Sheets no: con "=Listas!$A$4:$A$5" el desplegable no aparece.
    # Tope de 255 caracteres; la lista más larga (profesores) va en ~96.
    inline = ",".join(str(v) for v in valores)
    assert len(inline) <= 250, f"lista {titulo} muy larga para ir en línea ({len(inline)})"
    assert "," not in "".join(str(v) for v in valores), f"lista {titulo}: un valor trae coma"
    rangos[titulo] = f'"{inline}"'

# ═══════════════════ Hoja 2: Inscritos ═══════════════════
ws = wb.create_sheet("Inscritos", 1)
ws.freeze_panes = "A2"

borde = Border(bottom=Side(style="thin", color="CCCCCC"))
for i, (nombre, ancho, ayuda) in enumerate(COLUMNAS, start=1):
    L = get_column_letter(i)
    ws.column_dimensions[L].width = ancho
    c = ws.cell(row=1, column=i, value=nombre)
    c.font = Font(name=ARIAL, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.comment = Comment(ayuda, "Centro de Control CDAF", width=320, height=110)
ws.row_dimensions[1].height = 30

# Filas de ejemplo, en amarillo para que se vean y se borren
for r, fila in enumerate(EJEMPLOS, start=2):
    for c_idx, valor in enumerate(fila, start=1):
        c = ws.cell(row=r, column=c_idx, value=valor)
        c.font = Font(name=ARIAL, size=11)
        c.fill = PatternFill("solid", fgColor=AMARILLO)
        c.border = borde

# No se pinta estilo celda por celda en las filas vacías: crearía ~4.400 celdas con
# formato y triplica el peso del archivo sin cambiar nada a la vista. La hora se
# escribe como texto "16:30" y así la lee el importador.
FILAS_LIBRES = 400

ULTIMA = 1 + len(EJEMPLOS) + FILAS_LIBRES

# ── Listas desplegables ──
# OJO: en openpyxl `showDropDown=True` ESCONDE la flecha (el atributo del XML está
# invertido respecto a su nombre). Se deja sin tocar para que la flecha SÍ aparezca.
VALIDACIONES = [
    ("D", "Deporte",  "Escoge Tenis o Pádel de la lista."),
    ("E", "Academia", "Escoge Recreativa o Competencia de la lista."),
    ("F", "Nivel",    "Escoge el nivel del niño de la lista."),
    ("G", "Día",      "Escoge el día de la lista. Una fila por cada día que viene."),
    ("I", "Duración", "Escoge 60, 90 o 120 minutos."),
    ("J", "Profesor", "Escoge el profesor de la lista. Si falta alguien, dínoslo y lo agregamos."),
    ("K", "Cancha",   "Escoge el número de cancha de la lista."),
]
for col, titulo, mensaje in VALIDACIONES:
    dv = DataValidation(type="list", formula1=rangos[titulo], allow_blank=True)
    dv.error = mensaje
    dv.errorTitle = f"{titulo} no válido"
    dv.showErrorMessage = True
    dv.prompt = mensaje
    dv.promptTitle = titulo
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{ULTIMA}")

wb.save(SALIDA)
print("listo:", SALIDA)
print("hojas:", wb.sheetnames)
